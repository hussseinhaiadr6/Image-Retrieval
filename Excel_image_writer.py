from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import json
import os
from openpyxl.drawing.image import Image

def write_result_to_excel(excel_file,empty_column):
    excel_file=excel_file.replace('\\','/')
    txt_file = f"./extracted_{excel_file.split("/")[-1].split('.')[0]}.txt"
    # Load your workbook and sheet as you want, for example
    wb = load_workbook(excel_file)
    ws = wb.active
    with open(txt_file, 'r') as f:
        loaded_image_dict = json.load(f)

    for filename in os.listdir(f"./extracted_{excel_file.split("/")[-1].split('.')[0]}"):
        sheetname=filename.split("_")[0]
        print("for the image", filename)
        print("shhetname is",sheetname)
        row_nb=filename.split("_")[1].split('.')[0]
        print("row number is ", row_nb)
        sheet = wb[sheetname]
 
        row = sheet[row_nb]

        sheet.cell(row=int(row_nb), column=empty_column+1).value="_".join(loaded_image_dict[filename][0].split("_")[1:])
        sheet.cell(row=int(row_nb), column=empty_column + 2).value = "_".join(loaded_image_dict[filename][1].split("_")[1:])
        sheet.cell(row=int(row_nb), column=empty_column + 3).value = "_".join(loaded_image_dict[filename][2].split("_")[1:])
        


    wb.save(f"./{excel_file.split("/")[-1].split('.')[0]}_new.xlsx")








"""image_loader = SheetImageLoader(sheet)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row-1):

        for cell in row:
            if cell.column==2:
                # Put your sheet in the loader
                if cell.row>0 :

                    # And get image from specified cell
                    try:
                        image = image_loader.get('D'+str(cell.row))
                        image.save("./extracted_2/" + sheet_name + "_"+str(cell.row)+".png")
                    except:
                        break

                    # Image now is a Pillow image, so you can do the following


                    # Ask if there's an image in a cell
"""