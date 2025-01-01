from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import os


def Extract_images_from_excel(file_path,photo_column_name,product_column_name):
    # Load the workbook
    file_path=file_path.replace("\\","/")
    print(file_path)
    wb = load_workbook(file_path)

    # Create a directory to save the images if it doesn't exist
    output_dir = f"./extracted_{file_path.split("/")[-1].split('.')[0]}"
    os.makedirs(output_dir, exist_ok=True)

    # Iterate over all sheets in the workbook
    for sheet_name in wb.sheetnames:
        print(sheet_name)
        sheet = wb[sheet_name]
        image_loader = SheetImageLoader(sheet)

        # Find the column indexes for "PHOTO" and the product name (column "3")
        photo_col = None
        product_name_col = None

        for col in sheet.iter_cols(0, sheet.max_column):

            if col[0].value == photo_column_name:
                photo_col = col[0].column_letter
                print("photo col= ",photo_col)
            elif col[0].value == product_column_name:
                product_name_col = col[0].column_letter
                print("product_name col= ",product_name_col)
            if photo_col and product_name_col:
                break

        if not photo_col or not product_name_col:
            continue  # Skip this sheet if "PHOTO" or "3" column is not found

        # Iterate over the rows and extract the images
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row_num = row[0].row
            try:
                image = image_loader.get(photo_col + str(row_num))
                product_name = sheet[product_name_col + str(row_num)].value
                if "/"  in product_name:
                    product_name=product_name.replace("/","-")
                    print(product_name)

                if product_name:
                    image.save(f"{output_dir}/{sheet_name}_{row_num}_{product_name}.png")
            except:

                continue  # Skip if there is no image or any error occurs
    return output_dir
